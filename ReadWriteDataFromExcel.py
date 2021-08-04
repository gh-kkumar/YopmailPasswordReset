import openpyxl

def RowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_row)

def ColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return (sheet.max_column)

def FindColumnNoByName(FileName, SheetName, ColumnName):
    workbook = openpyxl.load_workbook(FileName)
    sheet = workbook.get_sheet_by_name(SheetName)
    MaxColumn = sheet.max_column
    RequiredColumn = 0

    for ColumnNo in range(MaxColumn):
        if(sheet.cell(row = 1, column = ColumnNo + 1).value == ColumnName):
            RequiredColumn = ColumnNo + 1
            break

    return RequiredColumn

def FindColumnNoByName1(FileName, SheetName, ColumnName, Row):
    workbook = openpyxl.load_workbook(FileName)
    sheet = workbook.get_sheet_by_name(SheetName)
    MaxColumn = sheet.max_column
    RequiredColumn = 0

    for ColumnNo in range(MaxColumn):
        if(sheet.cell(row = Row, column = ColumnNo + 1).value == ColumnName):
            RequiredColumn = ColumnNo + 1
            break

    return RequiredColumn

def ReadData(file, sheetname, rowno, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    return sheet.cell(row=rowno, column=columnno).value

def WriteData(file, sheetname, rowno, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetname)
    sheet.cell(row=rowno, column=columnno).value = data
    workbook.save(file)